from io import BytesIO  # ★ 新增
from openpyxl import load_workbook
from .a_extractInfo import ExcelReader
from . import z_tools
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import logging


logger = logging.getLogger('报关资料5')



class BaoGuanBuilder():
    def __init__(self, sourcepath= "./app/app_tasks/resource/baoguan.xlsx"):
        self.title = 'baoguan'
        self.wb = load_workbook(sourcepath)
        self.ws = self.wb.active
        self.ws.title = self.title
        self.thin_border = Border(
            left=Side(style='thin', color="000000"),
            right=Side(style='thin', color="000000"),
            top=Side(style='thin', color="000000"),
            bottom=Side(style='thin', color="000000"),
        )
        self.yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    def add_detial_info(self, itemInfo):
        total_jin = 0
        total_mao = 0
        total_num = 0
        for key, value in itemInfo.items():
            a, b, c = key
            heyue = value['预约号']
            fahuo = value['境内货源地']
            total_jin += value['净重']
            total_mao += value['毛重']
            total_num += value['箱数']
        logger.info(f'预约号：{heyue}，发货地：{fahuo}，总净重：{total_jin}，总毛重：{total_mao}，总箱数：{total_num}')
        self.ws.cell(row=16, column=1).value = heyue
        self.ws.cell(row=16, column=2).value = total_num
        self.ws.cell(row=16, column=7).value = total_mao
        self.ws.cell(row=16, column=11).value = total_jin
        self.ws.cell(row=12, column=10).value = fahuo

    def insertItem(self, itemInfo):
        content = [['项号   商品编号', 
                    "商品名称、规格型号", "", "", '',
                    '数量及单位',"",
                    '最终目的国（地区）', '','',
                    '单价/总价/币制  征免', '']]
        
        titlefont = Font(name="仿宋_GB2312", size=9)
        title_alignment = Alignment(horizontal="left", vertical="center")
        row_num = z_tools.record_sheet(self.ws, content, 22, 1, titlefont, title_alignment)
        z_tools.add_outer_border(self.ws, row_num-1, row_num-1, 1, 12)

        startIdex = 1
        for key, value in itemInfo.items():
            a, b, c = key
            showName = f' {startIdex:02d}  {a}'
            logger.info(showName)
            if c == '':
                c = '有问题！！！！'

            inputContent = [
                [showName, '', '', '', None, value['数量'], '条', '中国（CHN）', '', '', b, '照章征税'],
                [None, c, None, None, value['净重'], '', '千克', None, None, value['总价'], '', None],
                [None, None, None, None, None, None, None, None, None, None, "日本元", None]
            ]

            titlefont = Font(name="Times New Roman", size=10, bold=True)
            title_alignment = Alignment(horizontal="left", vertical="center")
            row_num = z_tools.record_sheet(self.ws, inputContent, row_num, 1, titlefont, title_alignment)
            merge_str = f'B{row_num-2}:D{row_num-1}'
            self.ws.merge_cells(merge_str)
            cell = self.ws.cell(row=row_num-2, column=2)
            cell.font = Font(name="Times New Roman", size=8, bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            rigtCellList = [(row_num-3, 6), (row_num-2, 6), (row_num-3, 11), (row_num-2, 10), (row_num-1, 11)]
            for rigtCell in rigtCellList:
                self.ws.cell(row=rigtCell[0], column=rigtCell[1]).alignment = Alignment(horizontal='right', vertical='center')

            self.ws.cell(row=row_num-3, column=12).alignment = Alignment(horizontal='center', vertical='center')

            self.ws.merge_cells(f'E{row_num-2}:F{row_num-2}')
            self.ws.cell(row=row_num-2, column=5).number_format = '0.00'
            self.ws.cell(row=row_num-2, column=5).alignment = Alignment(horizontal='right', vertical='center')

            z_tools.add_outer_border(self.ws, row_num-3, row_num-1, 1, 12)
            for rowIdx in range(row_num-3, row_num):
                self.ws.row_dimensions[rowIdx].height = 17

            startIdex += 1

        return row_num
    
    def addEndContent(self, row_num):
        weightList = [20.25, 13.5, 24.75, 14.25, 22.5, 
                      14.25, 21.75,14.25, 12, 9.75,
                      14.25, 14.25, 14.25, 14.25]

        for index in range(len(weightList)):
            self.ws.row_dimensions[row_num + index].height = weightList[index]

        content = [
            [" ", '', '', '', '', '', '', '', '', '', '', ''],
            ['特殊关系确认： ', '否', '价格影响确认：', '', '否', None, None, None, '支付特许权使用费确认：', '', '', '否'],
            ['报关员 ', None, None, None, None, None, None, ' 审单'],
            [None],
            ['单位地址', '', '申报单位（签章）', None, None, None, None, ' 征税'],
            [None],
            [None, None, None, None, None, None, None, ' 查验', None, None, '审价'],
            ['邮编              电话', '', '填制日期', ''],
            [None, None, None, None, None, None, None, None, None, None, '统计'],
            [None],
            [None, None, None, None, None, None, None, None, None, None, '放行'],
            [None],
            [None],
            [None, None, None, None, None, None, None, None, None, '海关编制', ''],
        ]
        font = Font(name="仿宋_GB2312", size=9)
        alignment = Alignment(horizontal="left", vertical="bottom")
        endRow = z_tools.record_sheet(self.ws, content, row_num, 1, font, alignment)
        endRow -= 1
        z_tools.add_outer_border(self.ws, row_num, endRow, 1, 12)
        z_tools.add_outer_border(self.ws, row_num + 4, row_num + 4, 1, 12)

        z_tools.add_outer_border(self.ws, row_num + 5, row_num + 5, 8, 12)
        z_tools.add_outer_border(self.ws, row_num + 6, row_num + 6, 8, 12)
        z_tools.add_outer_border(self.ws, row_num + 7, row_num + 7, 8, 12)
        z_tools.add_outer_border(self.ws, row_num + 9, row_num + 9, 8, 12)
        z_tools.add_outer_border(self.ws, row_num + 10, row_num + 10, 8, 12)

        z_tools.add_outer_border(self.ws, row_num + 6, row_num + 6, 1, 2, border_type='bottom')
        z_tools.add_outer_border(self.ws, row_num + 7, row_num + 7, 1, 2, border_type='bottom')
        z_tools.add_outer_border(self.ws, row_num + 8, row_num + 8, 1, 2, border_type='bottom')

    def save_data(self, save_path):
        self.wb.save(save_path)

    def to_bytes(self) -> bytes:
        """★ 新增：把当前工作簿写入内存并返回 bytes"""
        bio = BytesIO()
        self.wb.save(bio)
        return bio.getvalue()

    def process_data(self, extract_info):
        total_data = {}
        for item in extract_info:
            ywenstr = item.get('英文品名', '')
            zwenstr = item.get('中文品名', '')
            hsCode  = item.get('HS CODE', '')
            show_name = f"{hsCode}{zwenstr}"

            price       = item.get('单价', 0)
            amount      = item.get('总价', 0)
            jinWeight   = item.get('净重', 0)
            maoWeight   = item.get('毛重', 0)
            number      = item.get('数量', 0)
            xinghao     = item.get('产品型号', '')
            xiangshu    = item.get('箱数', 0)

            key = (show_name, price, xinghao)
            if key not in total_data:
                total_data[key] = {
                    '数量': number,
                    '总价': amount,
                    '净重': jinWeight,
                    '毛重': maoWeight,
                    '箱数': xiangshu,
                    '境内货源地': item.get('发货地', ''),
                    '预约号': item.get('合同号码', ''),
                }
            else:
                total_data[key]['数量'] += number
                total_data[key]['总价'] += amount
                total_data[key]['净重'] += jinWeight
                total_data[key]['毛重'] += maoWeight
                total_data[key]['箱数'] += xiangshu
        return total_data
    
    # ★ 改造点：返回 bytes；为兼容旧用法，save_path 设为可选
    def detect(self, dataDict, save_path: str | None = None) -> bytes:
        dataInfo = self.process_data(dataDict)
        rowNum = self.insertItem(dataInfo)
        self.addEndContent(rowNum)
        self.add_detial_info(dataInfo)
        if save_path:
            self.save_data(save_path)
        return self.to_bytes()


if __name__ == '__main__':
    save_path = '5-报关资料样板.xlsx'
    fp = r"app/app_tasks/resource/example.xlsx"
    with open(fp, "rb") as f:
        blob = f.read()
    reader = ExcelReader(blob)
    data_dicts = reader.read_as_dicts()

    builder = BaoGuanBuilder()
    xlsx_bytes = builder.detect(data_dicts, save_path=None)  # 只要 bytes 不落盘就传 None
    # 本地想看看文件的话：
    with open(save_path, 'wb') as f:
        f.write(xlsx_bytes)
