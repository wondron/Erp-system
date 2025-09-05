from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from enum import Enum

class Direction(Enum):
    HORIZONTAL = "horizontal"
    VERTICAL = "vertical"


def set_sheet_size(sheetobj, widthlist, heightlist):
    for idx, width in enumerate(widthlist, start=1):
        col_letter = get_column_letter(idx)
        sheetobj.column_dimensions[col_letter].width = width

    # 设置行高（row = 数字）
    for idx, height in enumerate(heightlist, start=1):
        sheetobj.row_dimensions[idx].height = height


def add_outer_border(ws, min_row, max_row, min_col, max_col, style="thin", color="000000", border_type="outer"):
    """
    给指定区域添加边框，可选择上下左右或整个外框
    ws: 工作表对象
    min_row, max_row: 区域起始行/结束行
    min_col, max_col: 区域起始列/结束列
    style: 边框样式，例如 "thin", "medium", "dashed"
    color: 边框颜色，十六进制字符串
    border_type: "top", "bottom", "left", "right", "outer"
    """
    side = Side(border_style=style, color=color)

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            
            top = side if (border_type in ["top", "outer"] and row == min_row) else None
            bottom = side if (border_type in ["bottom", "outer"] and row == max_row) else None
            left = side if (border_type in ["left", "outer"] and col == min_col) else None
            right = side if (border_type in ["right", "outer"] and col == max_col) else None
            
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)



def record_sheet(sheetobj, erweiData, start_row, start_col, fontConfig, alignment, konggeDirect=Direction.HORIZONTAL, border=None, color=None, isMerge=True):
    merged = set()  # 记录已处理过的单元格，避免重复合并
    maxRow = 0
    for r_offset, row in enumerate(erweiData):
        for c_offset, value in enumerate(row):
            r_idx = start_row + r_offset
            c_idx = start_col + c_offset
            if maxRow < r_idx:
                maxRow = r_idx
            if value == "":
                if not isMerge:
                    cell = sheetobj.cell(row=r_idx, column=c_idx, value=None)
                    cell.border = border
                merged.add((r_idx, c_idx))
                continue
            
            # 正常写入当前非空单元格
            cell = sheetobj.cell(row=r_idx, column=c_idx, value=value)
            cell.font = fontConfig
            cell.alignment = alignment
            if border:
                cell.border = border
            if color and value != 0:
                cell.fill = color

    if not isMerge:
        return maxRow + 1

    sorted_merged = sorted(merged, key=lambda x: (-x[0], -x[1]))
    used_item = []
    for kong_item in sorted_merged:
        idRow, idCol = kong_item
        if kong_item in used_item:
            continue

        if konggeDirect == Direction.HORIZONTAL:
            st_row = idRow
            ed_row = idRow
            test_col = idCol - 1
            while True:
                if sheetobj.cell(row=idRow, column=test_col).value != None and sheetobj.cell(row=idRow, column=test_col).value != "":
                    break
                used_item.append((idRow, test_col))
                test_col = test_col - 1
            # print(f"合并: R{st_row}C{test_col} -> R{ed_row}C{idCol}")
            sheetobj.merge_cells(start_row=st_row, start_column=test_col, end_row=ed_row, end_column=idCol)
        elif konggeDirect == Direction.VERTICAL:
            st_col = idCol
            ed_col = idCol
            test_row = idRow -1
            while True:
                if sheetobj.cell(row=test_row, column=st_col).value != None and sheetobj.cell(row=test_row, column=st_col).value != "":
                    break
                used_item.append((test_row, st_col))
                test_row = test_row - 1
            sheetobj.merge_cells(start_row=test_row, start_column=st_col, end_row=idRow, end_column=ed_col)
            # print(f"合并: R{test_row}C{st_col} -> R{idRow}C{ed_col}")
    return maxRow + 1


def change_data_type(format_dict, data):
    total_data = []

    for item in data:  # 每一行提取结果
        single_data = []
        for key, config in format_dict.items():
            has_val = False
            val = None
            # 遍历映射字段
            for val_item in config["yinshe"]:
                if val_item in item and item[val_item] not in [None, ""]:
                    val = item[val_item]
                    has_val = True
                    break

            # 如果没有取到，默认值为 0 或 ""
            if not has_val:
                if config["type"] == "str":
                    val = ""
                elif config["type"] == "int":
                    val = 0
                elif config["type"] == "float":
                    val = 0.0
            else:
                # 做类型转换
                try:
                    if config["type"] == "str":
                        val = str(val).strip()
                    elif config["type"] == "int":
                        val = int(float(val))  # 兼容小数字符串
                    elif config["type"] == "float":
                        val = float(val)
                except Exception:
                    # 转换失败兜底
                    if config["type"] == "str":
                        val = str(val)
                    elif config["type"] == "int":
                        val = 0
                    elif config["type"] == "float":
                        val = 0.0

            single_data.append(val)

        total_data.append(single_data)

    return total_data