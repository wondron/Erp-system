from __future__ import annotations
from io import BytesIO
from typing import List, Any, Union, Iterable, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image

# ✅ 统一用绝对导入，模块方式运行更稳
from app.app_tasks.baoGuan.z_tools import Direction
from app.app_tasks.baoGuan import z_tools
from app.app_tasks.baoGuan.a_extractInfo import ExcelReader


class ZhuangXiangBuilder:
    def __init__(self, title: str = "PackingList"):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = title

        self.thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        self.yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.row_height: List[float] = [33, 33, 16.1, 16.1, 16.1, 16.1, 16.1, 16.1, 16.1, 33, 16.1]
        self.col_weight: List[float] = [16, 39, 16, 16, 22, 22]

    # ---------- utils ----------
    @staticmethod
    def _to_number(v: Any) -> float:
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        if isinstance(v, str):
            s = v.strip().replace(",", "")
            if not s:
                return 0.0
            try:
                return float(s)
            except Exception:
                return 0.0
        return 0.0

    @staticmethod
    def _pad6(row: List[Any]) -> List[Any]:
        """把行补/裁到 6 列，防止写入时报错。"""
        return (row + [""] * 6)[:6]

    def _anchor_for_stamp(self, row_index: int) -> str:
        return f"E{max(1, row_index - 4)}"

    # ---------- layout ----------
    def change_sheet_size(self, max_row: int) -> None:
        max_row = max(1, int(max_row))
        if len(self.row_height) < max_row:
            need = max_row - len(self.row_height)
            if need > 0:
                if need >= 1:
                    self.row_height += [16.1] * (need - 2)
                    self.row_height += [33]
        z_tools.set_sheet_size(self.ws, self.col_weight, self.row_height)

    # ---------- content ----------
    def get_fix_content(self, heyuestr: str = " ") -> int:
        # 表头
        header1 = [["装 箱 单", "", "", "", "", ""]]
        header2 = [["PACKING LIST", "", "", "", "", ""]]

        # 卖方/买方信息
        content1 = [
            ["卖方：", "91330110MA28TYA536杭州同尘家居有限公司", ""],
            ["地址：", "浙江省杭州市余杭区余杭经济技术开发区北沙东路7号2幢324室", ""],
            ["买方：", " ", ""],
            ["地址：", " ", ""],
        ]

        content2 = [
            ["日期 Date:", " "],
            ["发票编号 Invoice No:", " "],
            ["合约号 Contract No:", heyuestr],
        ]

        # 船名 / 付款条件
        contente3 = [["船名", " ", "", "付款条件", " ", ""]]
        contente4 = [["Shipped by", " ", "", "Terms of Payment:", " ", ""]]

        # 表头列
        contente5 = [[
            "箱号\nCtn.No.",
            "货物名称及规格\nDescription",
            "总数(件)\nGe.Crate(CTNS)",
            "总数量\nGe.Quantity",
            "总毛重(千克)\nG.W.: (KG)",
            "总净重(千克)\nN.W.: (KG)",
        ]]

        # 中文/英文大标题
        font_title = Font(name="等线", size=22)
        center = Alignment(horizontal="center", vertical="center")
        rowindex = z_tools.record_sheet(self.ws, header1, 1, 1, font_title, center)

        font_title_en = Font(name="等线", size=20)
        rowindex = z_tools.record_sheet(self.ws, header2, rowindex, 1, font_title_en, center)

        # 卖方/买方信息
        font_small = Font(name="等线", size=10)
        left = Alignment(horizontal="left", vertical="center")
        _ = z_tools.record_sheet(self.ws, content1, rowindex, 1, font_small, left, border=self.thin_border)
        rowindex = z_tools.record_sheet(self.ws, content2, rowindex, 5, font_small, left, border=self.thin_border)

        # 船名 + 付款条件
        font_cn = Font(name="等线", size=12)
        left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
        rowindex = z_tools.record_sheet(self.ws, contente3, rowindex + 2, 1, font_cn, left_wrap)
        font_en = Font(name="等线", size=9)
        rowindex = z_tools.record_sheet(self.ws, contente4, rowindex, 1, font_en, left_wrap)
        # 外边框（两行）
        z_tools.add_outer_border(self.ws, rowindex - 2, rowindex - 1, 1, 6)
        z_tools.add_outer_border(self.ws, rowindex - 2, rowindex - 1, 1, 1)
        z_tools.add_outer_border(self.ws, rowindex - 2, rowindex - 1, 4, 4)

        # 列标题
        font_header = Font(name="等线", size=11)
        center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
        rowindex = z_tools.record_sheet(self.ws, contente5, rowindex, 1, font_header, center_wrap, border=self.thin_border)
        return rowindex

    def input_data(self, data: List[List[Any]], startRow: int) -> int:
        font = Font(name="微软雅黑", size=10)
        center = Alignment(horizontal="center", vertical="center")
        rowNum = z_tools.record_sheet(
            self.ws, data, startRow, 1, font, center,
            Direction.VERTICAL, border=self.thin_border, isMerge=False
        )
        return rowNum

    def get_zhuangxiang_data(self, extract_info: Iterable[Dict[str, Any]]) -> List[List[Any]]:
        """
        列顺序：
        0 箱号, 1 货物名称, 2 总数(件), 3 总数量, 4 总毛重, 5 总净重
        """
        show_data = {
            "箱号":   {"yinshe": ["ASIN"],   "type": "str"},
            "货物名称": {"yinshe": ["英文品名"], "type": "str"},
            "总数":   {"yinshe": ["箱数"],   "type": "int"},
            "总数量": {"yinshe": ["数量"],   "type": "int"},
            "总毛重": {"yinshe": ["毛重"],   "type": "float"},
            "总净重": {"yinshe": ["净重"],   "type": "float"},
        }
        table = z_tools.change_data_type(show_data, extract_info)

        cleaned: List[List[Any]] = []
        for row in table:
            # Packing List 不存在货币单位列，不填 JPY
            cleaned.append(self._pad6(list(row)))
        return cleaned

    def save_data(self, save_path: str) -> None:
        self.wb.save(save_path)

    def to_bytes(self) -> bytes:
        bio = BytesIO()
        self.wb.save(bio)
        return bio.getvalue()

    def calculate_total_data(self, start_row: int, data: List[List[Any]]) -> int:
        """
        合计列：2 总数(件)、3 总数量、4 总毛重、5 总净重
        """
        center = Alignment(horizontal="center", vertical="center")
        bold = Font(name="等线", size=12, bold=True)
        normal = Font(name="等线", size=12)
        wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if not data:
            total_row = [["合计 TOTAL", "", "0", "0", "0", "0"]]
            rowNum = z_tools.record_sheet(
                self.ws, total_row, start_row, 1, bold, center, Direction.HORIZONTAL, border=self.thin_border
            )
            self.ws.cell(row=rowNum - 1, column=1).alignment = Alignment(horizontal="right", vertical="center")

            marks = [["唛头\nMarks", "3V6S7", "", "", "", ""]]
            rowNum = z_tools.record_sheet(
                self.ws, marks, rowNum + 1, 1, normal, wrap_center, Direction.HORIZONTAL, border=self.thin_border
            )
            return rowNum

        sum_ctns = sum(self._to_number(r[2]) for r in data if len(r) > 2)  # 总数(件)
        sum_qty  = sum(self._to_number(r[3]) for r in data if len(r) > 3)  # 总数量
        sum_gw   = sum(self._to_number(r[4]) for r in data if len(r) > 4)  # 总毛重
        sum_nw   = sum(self._to_number(r[5]) for r in data if len(r) > 5)  # 总净重

        def _fmt_num(x: float) -> Any:
            return int(x) if abs(x - int(x)) < 1e-9 else round(x, 2)

        total_row = [["合计 TOTAL", "", _fmt_num(sum_ctns), _fmt_num(sum_qty), _fmt_num(sum_gw), _fmt_num(sum_nw)]]
        total_row[0] = self._pad6(total_row[0])

        rowNum = z_tools.record_sheet(
            self.ws, total_row, start_row, 1, bold, center, Direction.HORIZONTAL, border=self.thin_border
        )
        self.ws.cell(row=rowNum - 1, column=1).alignment = Alignment(horizontal="right", vertical="center")

        marks = [["唛头\nMarks", "3V6S7", "", "", "", ""]]
        rowNum = z_tools.record_sheet(
            self.ws, marks, rowNum + 1, 1, normal, wrap_center, Direction.HORIZONTAL, border=self.thin_border
        )
        return rowNum

    def insert_image(self, row_index: int, image_source: Union[str, bytes, bytearray, BytesIO]) -> None:
        if isinstance(image_source, (bytes, bytearray)):
            image_source = BytesIO(image_source)
        stamp = Image(image_source)
        stamp.width = 320
        stamp.height = 320
        self.ws.add_image(stamp, self._anchor_for_stamp(row_index))

    # ---------- main ----------
    def detect(self, oridata: Iterable[Dict[str, Any]], hetongstr: str, gongzhang_source: Optional[Union[str, bytes, bytearray, BytesIO]]) -> bytes:
        rowNum = self.get_fix_content(hetongstr)
        data = self.get_zhuangxiang_data(oridata)
        rowNum = self.input_data(data, rowNum)
        rowNum = self.calculate_total_data(rowNum, data)
        self.change_sheet_size(rowNum)
        if gongzhang_source:
            self.insert_image(rowNum, gongzhang_source)
        return self.to_bytes()


if __name__ == "__main__":
    # ✅ 从项目根目录 backend 运行模块方式：
    # python -m app.app_tasks.baoGuan.d_gen3file
    fp = r"app/app_tasks/resource/example.xlsx"

    # 方式1：文件路径
    # reader = ExcelReader(fp)
    # data_dicts = reader.read_as_dicts()

    # 方式2：bytes
    with open(fp, "rb") as f:
        blob = f.read()
    reader = ExcelReader(blob)
    data_dicts = reader.read_as_dicts()

    builder = ZhuangXiangBuilder()
    xlsx_bytes = builder.detect(
        data_dicts,
        data_dicts[0].get("合同号码", ""),
        gongzhang_source=r"app/app_tasks/resource/gonzhang.png",  # 也可传 bytes/BytesIO
    )

    with open("3-装箱单样板.xlsx", "wb") as f:
        f.write(xlsx_bytes)
