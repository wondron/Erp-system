from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from enum import Enum
from .z_tools import Direction
from . import z_tools
from typing import List, Any
from .a_extractInfo import ExcelReader


class ExcelASNBuilder:
    def __init__(self, title="ASN"):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = title

        # 默认样式：细边框、黄色填充
        self.thin_border = Border(
            left=Side(style='thin', color="000000"),
            right=Side(style='thin', color="000000"),
            top=Side(style='thin', color="000000"),
            bottom=Side(style='thin', color="000000"),
        )
        self.yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.row_height = [27.75, 21]
        self.col_weight = [9.88, 12.38, 21.23, 21.23, 12.73, 9.23, 9.23, 9.23, 10.63, 10.63, 11.28, 11.28, 5.52, 5.52, 5.52, 7.22]


    def change_sheet_size(self, max_row):
        if len(self.row_height) < max_row:
            self.row_height += (max_row - len(self.row_height)) * [21]
        aaa = self.row_height
        z_tools.set_sheet_size(self.ws, self.col_weight, self.row_height)

    
    def get_fix_content(self):
        # 2. 表头内容设置
        headers = [["Advance Shipping Notice", "", "", '']]
        content = [
            ["Vendor", "供货商"],
            ["Booking Key:", "", "提单号"],
            ["Appointment Key:", "", "预约号"],
            ["Truck information:", "", "车辆信息（车号 电话）"],
            ["Cargo delivery date:", "", "送货日期"]
        ]

        titlefont = Font(name="楷体", size=20, bold=True)
        title_alignment = Alignment(horizontal="center", vertical="center")
        contentfont = Font(name="楷体", size=10, bold=True, underline="single")
        content_alignment = Alignment(horizontal="left", vertical="center")

        rowindex = z_tools.record_sheet(self.ws, headers, 1, 1, titlefont, title_alignment)
        rowindex = z_tools.record_sheet(self.ws, content, rowindex, 1, contentfont, content_alignment)


        # 2. 数据标题行
        header_row_1 = [["PO#", 'SKU#', "中文品名", "英文品名", "HS CODE", "托数", "箱数", "件数", "单价", "总价", "净重", "毛重", "包装尺寸（m）", "", "", "体积"]]

        header_row_2 = [["",    "", "",     "",       "",           "pallet", "carton", "pc", "unit price", "amount", "", "", "长", "宽", "高", ""]]
        contentfont = Font(name="楷体", size=10, bold=True)
        rowindex = z_tools.record_sheet(self.ws, header_row_1, rowindex, 1, contentfont, title_alignment, border=self.thin_border)
        rowindex = z_tools.record_sheet(self.ws, header_row_2, rowindex, 1, contentfont, title_alignment, Direction.VERTICAL, border=self.thin_border)
        return rowindex

    def input_data(self, data, startRow):
        contentfont = Font(name="楷体", size=10)
        title_alignment = Alignment(horizontal="center", vertical="center")
        rowNum = z_tools.record_sheet(self.ws, data, startRow, 1, contentfont, title_alignment, Direction.VERTICAL, border=self.thin_border, isMerge=False)
        return rowNum

    def get_asn_data(self, extract_info):
        show_data = { 
            'PO#': {"yinshe": ['PO'], 'type': 'str'},
            'SKU#': {"yinshe": ['ASIN'], 'type': 'str'},
            '中文品名': {"yinshe": ['中文品名'], 'type': 'str'},
            '英文品名': {"yinshe": ['英文品名'], 'type': 'str'},
            'HS CODE': {"yinshe": ['海关编码'], 'type': 'str'},
            '托数': {"yinshe": ['托数'], 'type': 'int'},
            '箱数': {"yinshe": ['箱数'], 'type': 'int'},
            '件数': {"yinshe": ['数量'], 'type': 'int'},
            '单价': {"yinshe": ['单价'], 'type': 'int'},
            '总价': {"yinshe": ['总价'], 'type': 'int'},
            '净重': {"yinshe": ['净重'], 'type': 'float'},
            '毛重': {"yinshe": ['毛重'], 'type': 'float'},
            '长': {"yinshe": ['长'], 'type': 'int'},
            '宽': {"yinshe": ['宽'], 'type': 'int'},
            '高': {"yinshe": ['高'], 'type': 'int'},
            '体积': {"yinshe": ['体积'], 'type': 'float'}
        }

        total_data = []

        for item in extract_info:  # 每一行提取结果
            single_data = []
            for key, config in show_data.items():
                has_val = False
                val = None
                # 遍历映射字段
                for val_item in config["yinshe"]:
                    if val_item in item and item[val_item] not in [None, ""]:
                        val = item[val_item]
                        has_val = True
                        break

                # 如果没有取到，默认值为 0 或 ""
                if not has_val:
                    if config["type"] == "str":
                        val = ""
                    elif config["type"] == "int":
                        val = 0
                    elif config["type"] == "float":
                        val = 0.0
                else:
                    # 做类型转换
                    try:
                        if config["type"] == "str":
                            val = str(val).strip()
                        elif config["type"] == "int":
                            val = int(float(val))  # 兼容小数字符串
                        elif config["type"] == "float":
                            val = float(val)
                    except Exception:
                        # 转换失败兜底
                        if config["type"] == "str":
                            val = str(val)
                        elif config["type"] == "int":
                            val = 0
                        elif config["type"] == "float":
                            val = 0.0

                single_data.append(val)

            total_data.append(single_data)

        return total_data


    def save_data(self, save_path):
        # 保存文件
        self.wb.save(save_path)

    def calculate_total_data(self, start_row, data):
        show_data= {
            'PO#': False,
            'SKU#': False,
            '中文品名': False,
            '英文品名': False,
            '海关编码': False,
            '托数': ['托数'],
            '箱数': ['箱数'],
            '件数': ['数量'],
            '单价': False,
            '总价': ['总价'],
            '净重': ['净重'],
            '毛重': ['毛重'],
            '长': False,
            '宽': False,
            '高': False,
            '体积': ['体积'],
        }
        cal_index = []
        for index, item in enumerate(show_data):
            if show_data[item] != False:
                cal_index.append(index)
        if not data:
            return []

        n_cols = len(data[0])
        result = [''] * n_cols  # 其他列默认空字符串

        def to_number(s: str):
            """尝试把字符串转成 int 或 float，失败则返回 0"""
            try:
                if "." in s:   # 有小数点 → float
                    return float(s)
                return int(s)   # 否则 int
            except Exception:
                return 0

        for c in cal_index:
            total = 0.0
            for row in data:
                if c < len(row) and isinstance(row[c], str):
                    total += to_number(row[c])
                else:
                    total += row[c]
            # 判断是整数还是小数
            if abs(total - int(total)) < 1e-9:
                result[c] = int(total)  # 保持整数
            else:
                result[c] = round(total, 2)  # 保留两位小数
        result[0] = "TOTAL:"

        contentfont = Font(name="楷体", size=10, bold=True)
        title_alignment = Alignment(horizontal="center", vertical="center")
        rowNum = z_tools.record_sheet(self.ws, [result], start_row, 1, contentfont, title_alignment, Direction.VERTICAL, border=self.thin_border, isMerge=False)
        return rowNum


    def detect(self, oridata, save_path):
        asndata = self.get_asn_data(oridata)
        inputRow = self.get_fix_content()
        outputRow = self.input_data(asndata, inputRow)
        max_row = self.calculate_total_data(outputRow, asndata)
        self.change_sheet_size(max_row)
        self.save_data(save_path)



if __name__ == "__main__":
    save_path = '报关资料样板.xlsx'
    filepath = r'E:\01-code\02-erp\resource\DI信息.xlsx'
    reader = ExcelReader(filepath)
    data_dicts = reader.read_as_dicts()
    builder = ExcelASNBuilder()
    builder.detect(data_dicts, save_path)