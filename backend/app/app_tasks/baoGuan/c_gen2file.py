from __future__ import annotations
from io import BytesIO
from typing import List, Any, Union, Iterable, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image

# 建议：统一使用绝对导入，便于模块方式运行
from app.app_tasks.baoGuan.z_tools import Direction
from app.app_tasks.baoGuan import z_tools
from app.app_tasks.baoGuan.a_extractInfo import ExcelReader


class FaPiaoBuilder:
    def __init__(self, title: str = "fapiao"):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = title

        self.thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        self.yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        # 初始行高/列宽（后续会根据最大行数补齐）
        self.row_height: List[float] = [33, 33, 16.1, 16.1, 16.1, 16.1, 16.1, 33, 16.1]
        self.col_weight: List[float] = [16, 38.44, 14, 14, 22, 22]

    # ---------- utils ----------
    @staticmethod
    def _to_number(v: Any) -> float:
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        if isinstance(v, str):
            s = v.strip().replace(",", "")
            if not s:
                return 0.0
            try:
                return float(s)
            except Exception:
                return 0.0
        return 0.0

    @staticmethod
    def _pad6(row: List[Any]) -> List[Any]:
        """把行长度填充/裁剪到 6 列，防止写入时列数不齐。"""
        return (row + [""] * 6)[:6]

    def _anchor_for_stamp(self, row_index: int) -> str:
        # 盖章左上角锚点；可按需要调整偏移
        return f"E{max(1, row_index - 4)}"

    # ---------- layout ----------
    def change_sheet_size(self, max_row: int) -> None:
        """根据最大行数扩展行高列表，避免 set_sheet_size 越界。"""
        max_row = max(1, int(max_row))
        if len(self.row_height) < max_row:
            # 保留最后两个“加粗/尾部”行高：+…+33
            need = max_row - len(self.row_height)
            if need > 0:
                # 预留尾部 1 行 33 的空间；其余用 16.1 填充
                if need >= 1:
                    self.row_height += [16.1] * (need - 2)
                    self.row_height += [33]
        z_tools.set_sheet_size(self.ws, self.col_weight, self.row_height)

    # ---------- content ----------
    def get_fix_content(self, heyuestr: str = " ") -> int:
        header1 = [["发票", "", "", "", "", ""]]
        header2 = [["INVOICE", "", "", "", "", ""]]

        content1 = [
            ["卖方：", "91330110MA28TYA536杭州同尘家居有限公司", ""],
            ["地址：", "浙江省杭州市余杭区余杭经济技术开发区北沙东路7号2幢324室", ""],
            ["买方：", " ", ""],
            ["地址：", " ", ""],
        ]

        content2 = [
            ["日期 Date:", " "],
            ["发票编号 Invoice No:", " "],
            ["合约号 Contract No:", heyuestr],
        ]

        header3 = [[
            "箱号\nCtn.No.",
            "货物名称及规格\nDescription",
            "数量\nQuantity",
            "单位\nUnit",
            "单价\nUnit price",
            "总金额\nAmount",
        ]]

        # 标题
        titlefont = Font(name="等线", size=22)
        center = Alignment(horizontal="center", vertical="center")
        rowindex = z_tools.record_sheet(self.ws, header1, 1, 1, titlefont, center)

        # 英文标题
        contentfont = Font(name="等线", size=20)
        rowindex = z_tools.record_sheet(self.ws, header2, rowindex, 1, contentfont, center)

        # 卖方/买方信息
        smallfont = Font(name="等线", size=10)
        left = Alignment(horizontal="left", vertical="center")
        _ = z_tools.record_sheet(self.ws, content1, rowindex, 1, smallfont, left, border=self.thin_border)
        rowindex = z_tools.record_sheet(self.ws, content2, rowindex, 5, smallfont, left, border=self.thin_border)

        # 表头
        midfont = Font(name="等线", size=11)
        center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
        rowindex = z_tools.record_sheet(self.ws, header3, rowindex + 2, 1, midfont, center_wrap, border=self.thin_border)
        return rowindex

    def input_data(self, data: List[List[Any]], start_row: int) -> int:
        contentfont = Font(name="微软雅黑", size=10)
        center = Alignment(horizontal="center", vertical="center")
        rowNum = z_tools.record_sheet(
            self.ws, data, start_row, 1, contentfont, center,
            Direction.VERTICAL, border=self.thin_border, isMerge=False
        )
        return rowNum

    def get_fapiao_data(self, extract_info: Iterable[Dict[str, Any]]) -> List[List[Any]]:
        """
        将字典序列映射为二维数据：
        列顺序：箱号, 货物名称, 数量, 单位, 单价, 总数额
        """
        show_data = {
            "箱号":   {"yinshe": ["ASIN"],   "type": "str"},
            "货物名称": {"yinshe": ["英文品名"], "type": "str"},
            "数量":   {"yinshe": ["数量"],   "type": "int"},
            "单位":   {"yinshe": ["呵呵"],   "type": "str"},  # TODO: 替换为真实字段
            "单价":   {"yinshe": ["单价"],   "type": "int"},
            "总数额": {"yinshe": ["总价"],   "type": "int"},
        }
        table = z_tools.change_data_type(show_data, extract_info)

        cleaned: List[List[Any]] = []
        for row in table:
            row = list(row)
            # 单位为空 → 默认 'JPY'
            if len(row) > 3 and (row[3] is None or str(row[3]).strip() == ""):
                row[3] = "JPY"
            cleaned.append(self._pad6(row))
        return cleaned

    def save_data(self, save_path: str) -> None:
        self.wb.save(save_path)

    def to_bytes(self) -> bytes:
        bio = BytesIO()
        self.wb.save(bio)
        return bio.getvalue()

    def calculate_total_data(self, start_row: int, data: List[List[Any]]) -> int:
        """
        写入“合计 TOTAL”与“唛头 Marks”两行，并返回写入后的下一行行号。
        需要合计：数量(索引2)、总数额(索引5)
        """
        center = Alignment(horizontal="center", vertical="center")
        bold = Font(name="等线", size=12, bold=True)
        normal = Font(name="等线", size=12)
        wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if not data:
            # 没有数据也要写出结构，避免后续尺寸/盖章定位异常
            total_row = [["合计 TOTAL", "", "", " ", " ", "0"]]
            rowNum = z_tools.record_sheet(
                self.ws, total_row, start_row, 1, bold, center, Direction.HORIZONTAL, border=self.thin_border
            )
            self.ws.cell(row=rowNum - 1, column=1).alignment = Alignment(horizontal="right", vertical="center")

            marks = [["唛头\nMarks", "3V6S7", "", "", "", ""]]
            rowNum = z_tools.record_sheet(
                self.ws, marks, rowNum + 1, 1, normal, wrap_center, Direction.HORIZONTAL, border=self.thin_border
            )
            return rowNum

        # 合计数量/总数额
        sum_qty = sum(self._to_number(r[2]) for r in data if len(r) > 2)
        sum_amt = sum(self._to_number(r[5]) for r in data if len(r) > 5)
        total_row = [["合计 TOTAL", "", sum_qty, " ", " ", int(sum_amt) if abs(sum_amt - int(sum_amt)) < 1e-9 else round(sum_amt, 2)]]
        total_row[0] = self._pad6(total_row[0])

        rowNum = z_tools.record_sheet(
            self.ws, total_row, start_row, 1, bold, center, Direction.HORIZONTAL, border=self.thin_border
        )
        # “合计 TOTAL” 右对齐
        self.ws.cell(row=rowNum - 1, column=1).alignment = Alignment(horizontal="right", vertical="center")

        marks = [["唛头\nMarks", "3V6S7", "", "", "", ""]]
        rowNum = z_tools.record_sheet(
            self.ws, marks, rowNum + 1, 1, normal, wrap_center, Direction.HORIZONTAL, border=self.thin_border
        )
        return rowNum

    def insert_image(self, row_index: int, image_source: Union[str, bytes, bytearray, BytesIO]) -> None:
        if isinstance(image_source, (bytes, bytearray)):
            image_source = BytesIO(image_source)
        stamp = Image(image_source)  # str 路径或 BytesIO
        stamp.width = 320
        stamp.height = 320
        self.ws.add_image(stamp, self._anchor_for_stamp(row_index))

    # ---------- main ----------
    def detect(self, oridata: Iterable[Dict[str, Any]], hetongstr: str, gongzhang_source: Optional[Union[str, bytes, bytearray, BytesIO]]) -> bytes:
        rowNum = self.get_fix_content(hetongstr)
        total_data = self.get_fapiao_data(oridata)
        rowNum = self.input_data(total_data, rowNum)
        rowNum = self.calculate_total_data(rowNum, total_data)
        self.change_sheet_size(rowNum)
        if gongzhang_source:
            self.insert_image(rowNum, gongzhang_source)
        return self.to_bytes()


if __name__ == "__main__":
    # ✅ 建议从项目根目录（backend）以模块方式运行：
    # python -m app.app_tasks.baoGuan.c_gen2file
    # 下面示例：既支持“文件路径”，也支持“bytes”读入。
    # 1) 文件路径
    fp = r"app/app_tasks/resource/example.xlsx"
    with open(fp, "rb") as f:
        blob = f.read()
    reader = ExcelReader(blob)  # 传 bytes
    data_dicts = reader.read_as_dicts()

    # 2) bytes（如果用 UploadFile.read() 得到的原始二进制）
    # with open(fp, "rb") as f:
    #     blob = f.read()
    # reader = ExcelReader(blob)
    # data_dicts = reader.read_as_dicts()

    builder = FaPiaoBuilder()
    xlsx_bytes = builder.detect(
        data_dicts,
        data_dicts[0].get("合同号码", ""),
        gongzhang_source=r"app/app_tasks/resource/gonzhang.png",  # 也可传入 bytes/BytesIO
    )

    with open("2-报关资料样板.xlsx", "wb") as f:
        f.write(xlsx_bytes)
