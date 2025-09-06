from io import BytesIO  # ★ 新增
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from enum import Enum
from app.app_tasks.baoGuan.z_tools import Direction
from app.app_tasks.baoGuan import z_tools
from typing import List, Any
from .a_extractInfo import ExcelReader
from openpyxl.drawing.image import Image


class HeTongBuilder:
    def __init__(self, title="ZhangXiang"):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = title

        # 默认样式：细边框、黄色填充
        self.thin_border = Border(
            left=Side(style='thin', color="000000"),
            right=Side(style='thin', color="000000"),
            top=Side(style='thin', color="000000"),
            bottom=Side(style='thin', color="000000"),
        )
        self.yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.row_height = [52.5, 12, 13.5, 15.8, 13.5, 
                           15, 16.5, 13.5, 16.5, 13.5, 
                           16.5, 13.5, 13.5, 13.5, 13.5,
                           13.5, 15.5, 13.5,]
        self.col_weight = [20, 18.22, 13, 15, 13.22, 15.11, 4.11]


    def change_sheet_size(self, max_row):
        if len(self.row_height) < max_row:
            self.row_height += (max_row - len(self.row_height)-1) * [16.1]
        z_tools.set_sheet_size(self.ws, self.col_weight, self.row_height)

    
    def get_fix_content(self, heyuestr=" "):
        # 2. 表头内容设置
        header1 = [["合       同\nCONTRACT", "", "", '', '', '','']]

        content1 = [
            ["卖    方", "91330110MA28TYA536杭州同尘家居有限公司"],
            ["Sellers:", ],
            ["地    址", "浙江省杭州市余杭区余杭经济技术开发区北沙东路7号2幢324室"],
            ["Address:", " ", ' ', ' ', '预约号'],
            ['电    话', '15658813815', '传  真', ' ', 'Contract No:', heyuestr],
            ['TEL:', ' ', 'FAX:', ' ', '日     期'],
            ['买    方', ' ', ' ', ' ', 'Date:'],
            ['Buyers:', ' ', ' ', ' ', '签约地点'],
            ['地    址', ' ', ' ', ' ', 'Signed at:'],
            ['Address:', ' ', ' ', ' ', ' '],
            ['电    话:', ' ', '传  真'],
            ['TEL:', ' ', 'FAX:'],
        ]

        titlefont = Font(name="等线", size=22)
        title_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        rowindex = z_tools.record_sheet(self.ws, header1, 1, 1, titlefont, title_alignment)

        contentfont = Font(name="等线", size=11)
        title_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        rowindex = z_tools.record_sheet(self.ws, content1, 3, 1, contentfont, title_alignment,Direction.VERTICAL)
        self.ws.merge_cells("B3:D4")
        z_tools.add_outer_border(self.ws, 3, 4, 2, 4, border_type = 'bottom')
        self.ws.merge_cells("B5:D6")
        z_tools.add_outer_border(self.ws, 5, 6, 2, 4, border_type = 'bottom')
        self.ws.merge_cells("B7:B8")
        self.ws.merge_cells("D7:D8")
        z_tools.add_outer_border(self.ws, 7, 8, 2, 2, border_type = 'bottom')
        z_tools.add_outer_border(self.ws, 7, 8, 4, 4, border_type = 'bottom')
        self.ws.merge_cells("B9:D10")
        z_tools.add_outer_border(self.ws, 9, 10, 2, 4, border_type = 'bottom')
        self.ws.merge_cells("B11:D12")
        z_tools.add_outer_border(self.ws, 11, 12, 2, 4, border_type = 'bottom')
        self.ws.merge_cells("B13:B14")
        self.ws.merge_cells("D13:D14")
        z_tools.add_outer_border(self.ws, 13, 14, 2, 2, border_type = 'bottom')
        z_tools.add_outer_border(self.ws, 13, 14, 4, 4, border_type = 'bottom')
        self.ws.merge_cells("F7:G7")
        self.ws.merge_cells("F8:G8")
        self.ws.merge_cells("F9:G9")
        self.ws.merge_cells("F10:G10")
        z_tools.add_outer_border(self.ws, 7, 7, 6, 7, border_type = 'bottom')
        z_tools.add_outer_border(self.ws, 9, 9, 6, 7, border_type = 'bottom')
        z_tools.add_outer_border(self.ws, 11, 11, 6, 7, border_type = 'bottom')

        cell = self.ws.cell(row=3, column=2)  # E列 = 第5列
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        cell = self.ws.cell(row=5, column=2)  # E列 = 第5列
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        cell = self.ws.cell(row=7, column=2)  # E列 = 第5列
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        content2 = [
            ['经买卖双方确认根据下列条款订立本合同', ''],
            ['This contract is made out by the Selers and Buyers as per the following terms and conditions mutuilly confirmed:', '', '', '', '', ''],
        ]
        contentfont = Font(name="等线", size=10)
        title_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        rowNum = z_tools.record_sheet(self.ws, content2, 15, 1, contentfont, title_alignment)

        content3 = [
            ['(1) 货物名称及规格', '', '(2) 数 量', '(3) 单 位', '(4) 单 价', '(5) 金 额', ''],
            ['Name of commodity ', '', 'Quantity', 'Unit', 'Unit Price', 'Amount', '']
        ]
        contentfont = Font(name="等线", size=11)
        title_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        rowNum = z_tools.record_sheet(self.ws, content3, rowNum, 1, contentfont, title_alignment, border=self.thin_border)

        return rowNum


    def input_data(self, data, startRow):
        contentfont = Font(name="等线", size=11)
        title_alignment = Alignment(horizontal="center", vertical="center")

        changeRow = startRow
        for key, value in data.items():
            content_list = []
            content_list.append(key[0])
            content_list.append('')
            content_list.append(value['数量'])
            content_list.append(value['单位'])
            content_list.append(value['单价'])
            content_list.append(value['金额'])
            content_list.append("JPY")
            changeRow = z_tools.record_sheet(self.ws, [content_list], changeRow, 1, contentfont, title_alignment, border=self.thin_border)
        return changeRow
    

    def get_zhuangxiang_data(self, extract_info):
        """
        根据 extract_info 构建包装数据字典：
        - 相同 show_name 且单价相同的归为一类，数量和总金额累加
        - show_name 或单价不同视为不同类
        - 单价和金额格式化为千分位 + 四位小数
        """
        total_data = {}

        for item in extract_info:
            ywenstr = item.get('英文品名', '')
            zwenstr = item.get('中文品名', '')
            show_name = f"{ywenstr} {zwenstr}"

            # 单价和总价
            price = item.get('单价', 0)
            amount = item.get('总价', 0)

            # key 使用 (show_name, 单价) 来区分不同类
            key = (show_name, price)

            if key not in total_data:
                total_data[key] = {
                    '数量': item.get('数量', 0),
                    '单位': '条',
                    '单价': "{:,.4f}".format(price),
                    '金额': "{:,.2f}".format(amount)
                }
            else:
                # 累加数量和金额
                total_data[key]['数量'] += item.get('数量', 0)
                old_amount = float(total_data[key]['金额'].replace(',', ''))
                total_data[key]['金额'] = "{:,.2f}".format(old_amount + amount)

        return total_data


    def save_data(self, save_path):
        # 保存文件（保留原方法，按需使用）
        self.wb.save(save_path)

    def to_bytes(self) -> bytes:
        """★ 新增：把当前工作簿写入内存并返回 bytes"""
        bio = BytesIO()
        self.wb.save(bio)
        return bio.getvalue()

    def calculate_total_data(self, start_row, data):
        total_data = []
        total_num = 0
        total_money = 0

        for key, value in data.items():
            total_num += value['数量']
            total_money += int(float(value['金额'].replace(',', '')))

        contentfont = Font(name="等线", size=11)
        title_alignment = Alignment(horizontal="center", vertical="center")
        formatted = "{:,.2f}".format(total_money)
        total_data = [['Total', '', total_num, ' ', ' ', formatted, 'JPY']]
        rowNum = z_tools.record_sheet(self.ws, total_data, start_row, 1, contentfont, title_alignment, border=self.thin_border)

        cell = self.ws.cell(row=rowNum-1, column=1)
        cell.font = Font(name="等线", size=11, bold=True)
        cell.alignment = Alignment(horizontal="right", vertical="center")

        content = [
            ['数量及总值允许有  2    %的增减。', '', ''],
            [' 2  % more or less both in amount and quantity allowed.    ', '', '', '', ''],
            ['合同总值（大写）', '', '', '', ''],
            ['Total Value in Word:   ', '', '', '', ''],
            ['包装及唛头', '', '', '', ''],
            ['Packing and shipping Marks:   ', '', '', '', ''],
            ['装运期', '', '', '', ''],
            ['Time of Shipment:', '', '', '', ''],
            ['装运口岸和目的地', '', '', '', ''],
            ['Loading Port & Destination:', '', '', '', ''],
            ['付款条件  ', '', '', '', ''],
            ['Terms of Payment', '', '', '', ''],
            ['装运标记  ', '', '', '', ''],
            ['Shipping Marks:  ', '', '', '', ''],
            ['  ']
        ]
        contentfont = Font(name="等线", size=12)
        title_alignment = Alignment(horizontal="left", vertical="center")
        finalRow = z_tools.record_sheet(self.ws, content, rowNum + 1, 1, contentfont, title_alignment)

        rowIndex = [rowNum + 1, rowNum + 2, rowNum + 4, rowNum + 6, rowNum + 8, rowNum + 10, rowNum + 12, rowNum + 14]
        for item in rowIndex:
            cell = self.ws.cell(row=item, column=1)
            cell.font = Font(name="等线", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")

        data = [
            ['卖  方', '', ' ', '买  方', ''],
            ['  '],
            ['THE SELLERS', '', ' ', 'THE BUYERS', '']
        ]
        contentfont = Font(name="等线", size=12)
        title_alignment = Alignment(horizontal="center", vertical="center")
        finalRow = z_tools.record_sheet(self.ws, data, finalRow, 1, contentfont, title_alignment)
        return finalRow
    
    
    def insert_image(self, row_index, image_path):
        stamp = Image(image_path)
        stamp.width = 300
        stamp.height = 300
        posite = f"B{row_index-11}"
        self.ws.add_image(stamp, posite)


    # ★ 改造点：detect 返回 bytes，不再落盘，不再需要 save_path 参数
    def detect(self, oridata, hetongstr, gongzhang_path) -> bytes:
        rowNum = self.get_fix_content(hetongstr)
        zhuangxiang_data = self.get_zhuangxiang_data(oridata)
        rowNum1 = self.input_data(zhuangxiang_data, rowNum)
        rowNum2 = self.calculate_total_data(rowNum1, zhuangxiang_data)
        self.change_sheet_size(rowNum2)
        if gongzhang_path:
            self.insert_image(rowNum2, gongzhang_path)
        return self.to_bytes()  # ★ 返回二进制


if __name__ == "__main__":
    # 示例：本地调试时依然可以把返回的 bytes 写到磁盘看看效果
    fp = r"app/app_tasks/resource/example.xlsx"
    with open(fp, "rb") as f:
        blob = f.read()
    reader = ExcelReader(blob)
    data_dicts = reader.read_as_dicts()
    
    builder = HeTongBuilder()
    xlsx_bytes = builder.detect(data_dicts, data_dicts[0]['合同号码'], gongzhang_path="app/app_tasks/resource/gonzhang.png")

    with open('4-合同样板.xlsx', 'wb') as f:
        f.write(xlsx_bytes)
