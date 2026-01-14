# app/infrastructure/repositories_goods.py
from __future__ import annotations
import logging
import math
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, Tuple, Optional, Literal, Sequence, List
from sqlalchemy import select, func, asc, desc
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import selectinload
import unicodedata
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment  # 如需自动换行可用
from app.domain.models import GoodsIn
from app.infrastructure.orm_models import Goods, SupplyInfo, CustomsInfo, MaterialUsage, GoodsRaw
from sqlalchemy.exc import IntegrityError
from dataclasses import dataclass

from app.app_tasks.barcode_task.gen_barcode import build_barcode_pdf, LabelRow
from app.app_tasks.barcode_task.gen_xiangmai import build_carton_mark_pdf, xLabelRow
from fastapi import HTTPException
# === 新增：使用模板导出服务（pandas + openpyxl 由 goods_outporter 负责） ===
from io import BytesIO
from app.infrastructure.services.goods_outporter import (
    export_from_goods_list,   # (goods_list -> bytes)
)

logger = logging.getLogger('repo.goods')


def _normalize_like_kw(s: str) -> str:
    # LIKE 中 % 和 _ 有特殊含义，做一下转义更稳（避免用户输入带通配符导致“误扩大”）
    return s.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")


async def export_goods_xlsx_all(
    session: AsyncSession,
    *,
    sheet_name: str = "Sheet1",
) -> BytesIO:
    """
    全量导出当前所有商品为 xlsx（含 rels）。
    注意：数据量很大时会占内存；一般 ERP 商品量可接受。
    """
    stmt = select(Goods).options(*_with_rels()).order_by(asc(Goods.id))
    res = await session.execute(stmt)
    goods_list = res.scalars().unique().all()

    base_data: bytes = export_from_goods_list(goods_list, sheet_name=sheet_name)
    tuned: bytes = _apply_autofit_on_xlsx_bytes(
        base_data,
        sheet_name=sheet_name,
        padding=2.0,
        min_width=8.0,
        max_width=100.0,
        wrap_long_text=False,
    )
    return BytesIO(tuned)


# === 自动适配列宽工具 ===
def _east_asian_len(s: str) -> int:
    """按 East Asian Width 估算显示宽度：全角(W/F)=2，其他=1。"""
    total = 0
    for ch in s:
        ea = unicodedata.east_asian_width(ch)
        total += 2 if ea in ("W", "F") else 1
    return total

def _best_line_len(text: str) -> int:
    """支持多行：返回各行中最大显示宽度。"""
    if not text:
        return 0
    lines = text.splitlines() or [text]
    return max(_east_asian_len(line) for line in lines)


def _autofit_columns_on_ws(ws, *, padding: float = 2.0, min_width: float = 8.0, max_width: float = 100.0,
                           wrap_long_text: bool = False) -> None:
    """
    遍历工作表所有列，按内容计算合适列宽并设置。
    - padding: 额外留白
    - min/max_width: 列宽上下限，避免过窄/过宽
    - wrap_long_text: True 时给所有单元格开启换行（若你更偏向“显示完整但可能增高行高”，可打开）
    """
    # 先可选地打开自动换行（在列宽受限时更易完整显示）
    if wrap_long_text:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

    # 计算每列最大宽度需求
    for col_idx in range(1, ws.max_column + 1):
        maxlen = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            val = row[0]
            if val is None:
                continue
            # 数字/日期等统一转字符串估算显示宽度
            text = str(val)
            maxlen = max(maxlen, _best_line_len(text))

        # openpyxl 的 width 是近似“字符数”概念
        width = max(min_width, min(maxlen + padding, max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        

def _apply_autofit_on_xlsx_bytes(data: bytes, *, sheet_name: str | None = None,
                                 **kwargs) -> bytes:
    """
    把 bytes 载入为工作簿，执行自动列宽，返回新的 bytes。
    kwargs 透传给 _autofit_columns_on_ws（如 padding/min_width/max_width/wrap_long_text）
    """
    bio = BytesIO(data)
    wb = load_workbook(bio)
    ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active

    _autofit_columns_on_ws(ws, **kwargs)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()

# ------------------------------ 序列化给前端 ------------------------------
def serialize_goods(g: Goods) -> dict:
    return {
        "id": g.id,
        "category": g.category,
        "subcategory": g.subcategory,
        "season": g.season,
        "product_name": g.product_name,
        "channel": g.channel,
        "owner": g.owner,
        "sku": g.sku,
        "asin": g.asin,
        "barcode": g.barcode,
        "carton_mark": g.carton_mark,
        'item_no': g.item_no,
        "color": g.color,
        "size": g.size,
        "sale_price": float(g.sale_price) if g.sale_price is not None else None,
        "supply": {
            "vendor": g.supply.vendor if g.supply else None,
            "purchase_price": float(g.supply.purchase_price) if (g.supply and g.supply.purchase_price is not None) else None,
            "pkg_size": g.supply.pkg_size if g.supply else None,
            "pkg_weight": float(g.supply.pkg_weight) if (g.supply and g.supply.pkg_weight is not None) else None,
            "packing_ratio": g.supply.packing_ratio if g.supply else None,
            "carton_l": g.supply.carton_l if g.supply else None,
            "carton_w": g.supply.carton_w if g.supply else None,
            "carton_h": g.supply.carton_h if g.supply else None,
        },
        "customs": {
            "name_cn": g.customs.name_cn if g.customs else None,
            "name_en": g.customs.name_en if g.customs else None,
            "hscode": g.customs.hscode if g.customs else None,
            "declaration": g.customs.declaration if g.customs else None,
            "declared_price": float(g.customs.declared_price) if (g.customs and g.customs.declared_price is not None) else None,
            "image_note": g.customs.image_note if g.customs else None,
        },
        "materials": [
            {"name": m.material_name, "qty": float(m.quantity), "unit": getattr(m, "unit", None) or "件"}
            for m in (g.materials or [])
        ],
    }

def _to_num(x) -> Decimal:
    try:
        return Decimal(str(x).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return Decimal(str(float(x)))

# ------------------------------ JSON 清洗 ------------------------------
def _json_sanitize(obj: Any):
    """
    递归清洗：
    - Decimal -> float（失败则转 str）
    - float 的 NaN/Inf -> None
    - dict/list/tuple 递归处理
    其余类型原样返回
    """
    if obj is None or isinstance(obj, (bool, int, str)):
        return obj

    if isinstance(obj, Decimal):
        try:
            return float(obj)
        except Exception:
            return str(obj)

    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj

    if isinstance(obj, dict):
        return {k: _json_sanitize(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_json_sanitize(v) for v in obj]
    if isinstance(obj, tuple):
        return tuple(_json_sanitize(v) for v in obj)

    return obj


# 统一的查询 options（带出一对一/一对多）
def _with_rels():
    return (
        selectinload(Goods.supply),
        selectinload(Goods.customs),
        selectinload(Goods.materials),
    )


def _extract_material_items(p: Any) -> List[Tuple[str, Decimal, str]]:
    root = getattr(p, "__root__", None)
    if root is None:
        root = getattr(p, "root", None)
    if root is None:
        root = p

    out: List[Tuple[str, Decimal, str]] = []

    if isinstance(root, list):
        for item in root:
            if isinstance(item, dict):
                name = item.get("name", None)
                if name is None:
                    name = item.get("材料", None)

                qty = item.get("qty", None)
                if qty is None:
                    qty = item.get("用量", None)

                unit = item.get("unit", None) or item.get("用量单位", None) or item.get("单位", None)

                if name is not None and qty is not None:
                    out.append((str(name), _to_num(qty), str(unit).strip() if unit else "件"))
            else:
                name = getattr(item, "name", None)
                qty  = getattr(item, "qty", None)
                unit = getattr(item, "unit", None)
                if name is not None and qty is not None:
                    out.append((str(name), _to_num(qty), str(unit).strip() if unit else "件"))
        return out

    if isinstance(root, dict):
        any_pair = False
        for k, v in root.items():
            ks = str(k)
            if ks.startswith("材料") and not ks.endswith("用量") and not ks.endswith("用量单位"):
                idx = ks.replace("材料", "")
                qty  = root.get(f"材料{idx}用量")
                unit = root.get(f"材料{idx}用量单位")
                if v is not None and qty is not None:
                    out.append((str(v), _to_num(qty), str(unit).strip() if unit else "件"))
                    any_pair = True
        if any_pair:
            return out

        name = root.get("name") or root.get("材料")
        qty  = root.get("qty")  or root.get("用量")
        unit = root.get("unit") or root.get("用量单位") or root.get("单位")
        if name is not None and qty is not None:
            out.append((str(name), _to_num(qty), str(unit).strip() if unit else "件"))
        return out

    return out


# ----------------------------- 仓储 API ------------------------------
async def list_goods_with_count(
    session: AsyncSession,
    *,
    offset: int = 0,
    limit: int = 100,
    order_by: str = "id",
    order: str = "desc",
    barcode_contains: Optional[str] = None,
) -> tuple[list[Goods], int]:
    """分页 + 总数 + 排序 + 条形码包含过滤（连续子串）"""

    # --- 安全地映射字段名，防止 SQL 注入 ---
    valid_columns = {c.name for c in Goods.__table__.columns}
    if order_by not in valid_columns:
        order_by = "id"
    order_func = desc if order.lower() == "desc" else asc

    # --- 构造统一过滤条件（total 和 items 共用）---
    where_clauses = []
    if barcode_contains:
        kw = _normalize_like_kw(barcode_contains.strip())
        if kw:
            # 连续子串：LIKE %kw%
            # escape='\\' 对应上面的转义
            where_clauses.append(Goods.barcode.like(f"%{kw}%", escape="\\"))

    # --- 总数查询（带过滤）---
    total_stmt = select(func.count()).select_from(Goods)
    for w in where_clauses:
        total_stmt = total_stmt.where(w)
    total_res = await session.execute(total_stmt)
    total = total_res.scalar_one()

    # --- 主查询（带过滤）---
    stmt = (
        select(Goods)
        .options(*_with_rels())
    )
    for w in where_clauses:
        stmt = stmt.where(w)

    stmt = (
        stmt.order_by(order_func(getattr(Goods, order_by)))
        .offset(offset)
        .limit(limit)
    )

    res = await session.execute(stmt)
    items = res.scalars().unique().all()
    return items, total


async def list_goods(session: AsyncSession, *, offset: int = 0, limit: int = 100) -> list[Goods]:
    stmt = select(Goods).options(*_with_rels()).offset(offset).limit(limit)
    res = await session.execute(stmt)
    return res.scalars().unique().all()


async def get_goods_by_barcodes(session: AsyncSession, barcodes: Sequence[str]) -> list[Goods]:
    if not barcodes:
        return []
    # 去重 & 清理
    uniq: list[str] = []
    seen = set()
    for b in barcodes:
        if not b:
            continue
        s = str(b).strip()
        if s and s not in seen:
            uniq.append(s)
            seen.add(s)
    if not uniq:
        return []
    stmt = select(Goods).where(Goods.barcode.in_(uniq)).options(*_with_rels())
    res = await session.execute(stmt)
    return res.scalars().unique().all()


async def delete_goods_by_barcodes(session: AsyncSession, barcodes: Sequence[str]) -> dict:
    """
    批量按产品条码删除；为保证关系级联，逐条 ORM 删除（不要用 bulk delete）。
    返回: {"deleted": [...], "not_found": [...], "count": N}
    """
    if not barcodes:
        return {"deleted": [], "not_found": [], "count": 0}

    # 去重并清理空值，同时保留顺序
    seen = set()
    uniq = []
    for b in barcodes:
        if b and b not in seen:
            uniq.append(b)
            seen.add(b)

    res = await session.execute(select(Goods).where(Goods.barcode.in_(uniq)))
    goods = res.scalars().all()
    found_set = {g.barcode for g in goods if g.barcode is not None}

    # 逐条删除以触发关系 cascade
    for g in goods:
        await session.delete(g)
    await session.flush()  # 交给 get_db()/get_session 统一 commit

    deleted = [b for b in uniq if b in found_set]
    not_found = [b for b in uniq if b not in found_set]
    return {"deleted": deleted, "not_found": not_found, "count": len(deleted)}


async def list_all_barcodes(session: AsyncSession) -> list[str]:
    """
    获取所有商品的条码列表（去重、去空、按条码升序）。
    返回：["barcode1", "barcode2", ...]
    """
    stmt = (
        select(Goods.barcode)
        .where(Goods.barcode.isnot(None))
        .where(Goods.barcode != "")
        .distinct()
        .order_by(asc(Goods.barcode))
    )
    res = await session.execute(stmt)
    # res.scalars().all() -> list[str | None]
    barcodes = [b for b in res.scalars().all() if b]
    return barcodes


async def delete_goods_by_barcode(session: AsyncSession, barcode: str) -> bool:
    res = await session.execute(select(Goods).where(Goods.barcode == barcode))
    g = res.scalar_one_or_none()
    if not g:
        return False
    await session.delete(g)  # 依赖 ORM 关系的 cascade 或 FK ondelete
    await session.flush()
    return True


async def update_goods_by_barcode(session: AsyncSession, barcode: str, update_payload) -> Goods | None:
    q = await session.execute(
        select(Goods).where(Goods.barcode == barcode).options(*_with_rels())
    )
    g: Optional[Goods] = q.scalar_one_or_none()
    if not g:
        return None

    # ---------- 销售信息 ----------
    s = getattr(update_payload, "销售信息", None)
    if s:
        if getattr(s, "SKU", None) is not None: g.sku = s.SKU
        if getattr(s, "ASIN", None) is not None: g.asin = s.ASIN
        if getattr(s, "产品", None) is not None: g.product_name = s.产品
        if getattr(s, "大类目", None) is not None: g.category = s.大类目
        if getattr(s, "小品类", None) is not None: g.subcategory = s.小品类
        if getattr(s, "季节性", None) is not None: g.season = s.季节性
        if getattr(s, "销售渠道", None) is not None: g.channel = s.销售渠道
        if getattr(s, "责任人", None) is not None: g.owner = s.责任人
        if getattr(s, "颜色", None) is not None: g.color = s.颜色
        if getattr(s, "尺寸", None) is not None: g.size = s.尺寸
        if getattr(s, "销售价", None) is not None: g.sale_price = s.销售价
        if getattr(s, "自定义箱唛", None) is not None: g.carton_mark = s.自定义箱唛
        if getattr(s, "货号", None) is not None: g.item_no = s.货号
        # 可选择允许改条码（注意唯一约束冲突）
        new_barcode = getattr(s, "产品条码", None)
        if new_barcode is not None and new_barcode != g.barcode:
            g.barcode = new_barcode

    # ---------- 供应信息 ----------
    sup = getattr(update_payload, "供应信息", None)
    if sup:
        if not g.supply:
            g.supply = SupplyInfo()
        if getattr(sup, "供应商", None) is not None: g.supply.vendor = sup.供应商
        if getattr(sup, "采购价", None) is not None: g.supply.purchase_price = sup.采购价
        if getattr(sup, "单品包装尺寸", None) is not None: g.supply.pkg_size = sup.单品包装尺寸
        if getattr(sup, "单品包装重量", None) is not None: g.supply.pkg_weight = sup.单品包装重量
        if getattr(sup, "装箱系数", None) is not None: g.supply.packing_ratio = sup.装箱系数
        if getattr(sup, "外箱长", None) is not None: g.supply.carton_l = sup.外箱长
        if getattr(sup, "外箱宽", None) is not None: g.supply.carton_w = sup.外箱宽
        if getattr(sup, "外箱高", None) is not None: g.supply.carton_h = sup.外箱高

    # ---------- 报关信息 ----------
    cs = getattr(update_payload, "报关信息", None)
    if cs:
        if not g.customs:
            g.customs = CustomsInfo()
        if getattr(cs, "中文品名", None) is not None: g.customs.name_cn = cs.中文品名
        if getattr(cs, "英文品名", None) is not None: g.customs.name_en = cs.英文品名
        if getattr(cs, "海关编码", None) is not None: g.customs.hscode = cs.海关编码
        if getattr(cs, "申报要素", None) is not None: g.customs.declaration = cs.申报要素
        if getattr(cs, "申报价", None) is not None: g.customs.declared_price = cs.申报价
        if getattr(cs, "图片", None) is not None: g.customs.image_note = cs.图片

    # ---------- 生产配套（材料） ----------
    prod = getattr(update_payload, "生产配套", None)
    replace_materials = getattr(update_payload, "replace_materials", True)
    if prod is not None:
        items = _extract_material_items(prod)

        if replace_materials:
            for m in list(g.materials or []):
                await session.delete(m)
            g.materials = []

        for name, qty, unit in items:
            g.materials.append(MaterialUsage(material_name=name, quantity=qty, unit=unit))

    # ---------- flush ----------
    try:
        await session.flush()
    except IntegrityError:
        # 常见：条码改成已有值，触发 uq_goods_barcode 冲突
        # 交给上层转换为 409 或返回友好错误
        raise

    return g



async def create_goods(session: AsyncSession, data: GoodsIn) -> Goods:
    """
    新增商品：
    - 写 goods_raw(JSONB)
    - 写结构化表 goods / supply / customs / materials
    - 不 commit（交给 get_db 统一提交）
    """
    s = data.销售信息

    logger.info(
        "create_goods: sku=%s, barcode=%s, asin=%s, product=%s",
        s.SKU, s.产品条码, s.ASIN, s.产品
    )

    # --- 原始 JSON（消毒后） ---
    raw_payload = _json_sanitize(data.model_dump(mode="python"))
    raw = GoodsRaw(payload=raw_payload)
    session.add(raw)
    await session.flush()  # 得到 raw.id

    # --- 主表 ---
    g = Goods(
        sku=s.SKU,
        asin=s.ASIN,
        product_name=s.产品,
        category=s.大类目,
        subcategory=s.小品类,
        season=s.季节性,
        channel=s.销售渠道,
        owner=s.责任人,
        color=s.颜色,
        size=s.尺寸,
        sale_price=s.销售价,
        barcode=s.产品条码,
        carton_mark=s.自定义箱唛,
        item_no=s.货号,
    )
    session.add(g)

    try:
        await session.flush()  # ✅ 在这里触发唯一约束检查 + 得到 g.id
    except IntegrityError as e:
        logger.warning(
            "create_goods 唯一性冲突：sku=%s, barcode=%s, asin=%s, err=%s",
            s.SKU, s.产品条码, s.ASIN, str(e),
            exc_info=True,
        )
        raise

    # 回填 raw.goods_id
    raw.goods_id = g.id

    # --- 供应信息（可选） ---
    sup = getattr(data, "供应信息", None)
    if sup is not None:
        session.add(SupplyInfo(
            goods_id=g.id,
            vendor=sup.供应商,
            purchase_price=sup.采购价,
            pkg_size=sup.单品包装尺寸,
            pkg_weight=sup.单品包装重量,
            packing_ratio=sup.装箱系数,
            carton_l=sup.外箱长,
            carton_w=sup.外箱宽,
            carton_h=sup.外箱高,
        ))

    # --- 报关信息（可选） ---
    cs = getattr(data, "报关信息", None)
    if cs is not None:
        session.add(CustomsInfo(
            goods_id=g.id,
            name_cn=cs.中文品名,
            name_en=cs.英文品名,
            hscode=cs.海关编码,
            declaration=cs.申报要素,
            declared_price=cs.申报价,
            image_note=cs.图片,
        ))

    # --- 生产配套（可选） ---
    prod = getattr(data, "生产配套", None)
    for name, qty, unit in _extract_material_items(prod):
        session.add(MaterialUsage(
            goods_id=g.id,
            material_name=name,
            quantity=qty,
            unit=unit or "件",
        ))

    await session.flush()
    await session.refresh(g)
    return g

async def get_goods_by_sku(session: AsyncSession, sku: str) -> Goods | None:
    res = await session.execute(select(Goods).where(Goods.sku == sku))
    return res.scalar_one_or_none()



#------------------------------- 导出 PDF --------------------------------
async def export_goods_pdf(session: AsyncSession, barcode: str) -> BytesIO:
    rows = await get_goods_by_barcodes(session, [barcode])
    if not rows:
        # 用 HTTPException 让 FastAPI 正确返回 404，而不是让中间件捕获 ValueError
        raise HTTPException(status_code=404, detail=f"未找到条码 {barcode} 对应的商品")

    g = rows[0]
    info = LabelRow(g.color or "", g.size or "", g.barcode or "")

    # ✅ 改成位置参数；若函数只接收一个参数，就去掉 False
    pdf_bytes = build_barcode_pdf([info] * 24, False)

    buf = BytesIO(pdf_bytes)
    buf.seek(0)  # 保险起见，确保从头开始读
    return buf
    
    

async def export_carton_pdf(session: AsyncSession, barcode: str) -> BytesIO:
    goods_list = await get_goods_by_barcodes(session, [barcode])
    if not goods_list:
        raise HTTPException(status_code=404, detail=f"未找到条码 {barcode} 对应的商品")

    g = goods_list[0]
    info = xLabelRow(g.carton_mark or "", g.color or "", g.size or "", g.barcode or "")

    # 每页 6 个箱唛
    print_data = [info] * 6

    # ⚠️ 根据实际函数签名来决定是否传第二个参数
    pdf_bytes = build_carton_mark_pdf(print_data)   # 如果定义里只有 print_data

    buf = BytesIO(pdf_bytes)
    buf.seek(0)
    return buf


async def export_labels_pdf(
    session: AsyncSession,
    items: Dict[str, int],
    label_type: Literal["barcode", "carton_mark"],
) -> BytesIO:
    barcode_items = []
    logger.info(f'调用export_labels_pdf，items:{items}, type:{label_type}')
    if label_type == 'barcode':
        for barcode, count in items.items():
            goods_list = await get_goods_by_barcodes(session, [barcode])
            g = goods_list[0]
            info = LabelRow(g.color or "", g.size or "", g.barcode or "")
            logger.info(f'条码信息：{info}')
            barcode_items.extend([info] * count)
        pdf_bytes = build_barcode_pdf(barcode_items, False)
    elif label_type == 'carton_mark':
        for barcode, count in items.items():
            goods_list = await get_goods_by_barcodes(session, [barcode])
            g = goods_list[0]
            info = xLabelRow(g.carton_mark or "", g.color or "", g.size or "", g.barcode or "")
            logger.info(f'箱唛信息：{info}')
            barcode_items.extend([info] * count)
        pdf_bytes = build_carton_mark_pdf(barcode_items, False)
    
    buf = BytesIO(pdf_bytes)
    buf.seek(0)
    return buf


# ------------------------------ 导出 Excel ------------------------------
async def export_goods_xlsx_by_barcodes(
    session: AsyncSession,
    barcodes: Sequence[str],
    *,
    sheet_name: str = "Sheet1",
) -> BytesIO:
    goods_list = await get_goods_by_barcodes(session, barcodes)  # 先查
    base_data: bytes = export_from_goods_list(goods_list, sheet_name=sheet_name)
    tuned: bytes = _apply_autofit_on_xlsx_bytes(
        base_data,
        sheet_name=sheet_name,
        padding=2.0,
        min_width=8.0,
        max_width=100.0,
        wrap_long_text=False,
    )
    return BytesIO(tuned)