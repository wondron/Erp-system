# app/infrastructure/services/goods_outporter.py
from __future__ import annotations
import io
from typing import Any, Dict, Iterable, List, Tuple
from datetime import datetime, date, time, timezone
from decimal import Decimal
from uuid import UUID
from enum import Enum
from openpyxl.styles import Font, Alignment, Border, Side
import pandas as pd
from app.infrastructure.orm_models import Goods


# 参考 1111.xlsx (Sheet1) 的列顺序（首列选择框 + 销售/供应/报关 + 材料1..10及用量）
ORDERED_COLUMNS: List[str] = [
    # 选择/销售
    "选择框", "大类目", "小品类", "季节性", "产品", "销售渠道", "责任人",
    "SKU", "ASIN", "产品条码", "自定义箱唛", "货号", "颜色", "尺寸", "销售价",
    # 供应
    "供应商", "采购价", "单品包装尺寸", "单品包装重量", "装箱系数", "外箱长", "外箱宽", "外箱高",
    # 报关
    "中文品名", "英文品名", "海关编码", "申报要素", "申报价", "图片",
    # 材料（最多 1~10）
    "材料1", "材料1用量", "材料1用量单位",
    "材料2", "材料2用量", "材料2用量单位",
    "材料3", "材料3用量", "材料3用量单位",
    "材料4", "材料4用量", "材料4用量单位",
    "材料5", "材料5用量", "材料5用量单位",
    "材料6", "材料6用量", "材料6用量单位",
    "材料7", "材料7用量", "材料7用量单位",
    "材料8", "材料8用量", "材料8用量单位",
    "材料9", "材料9用量", "材料9用量单位",
    "材料10", "材料10用量", "材料10用量单位",
]

# 供路由调用的仓储方法

# -------------------- Excel 友好清洗 --------------------
def _excel_sanitize_value(val: Any):
    """
    - datetime/date/time: 去 tzinfo（必要时先转 UTC）；
    - Decimal -> float（失败则 str）；
    - Enum -> .value；
    - UUID/bytes/dict/list/tuple/set -> str；
    - 其余原样或 str。
    """
    if val is None or isinstance(val, (int, float, bool, str)):
        return val

    if isinstance(val, Decimal):
        try:
            return float(val)
        except Exception:
            return str(val)

    if isinstance(val, datetime):
        if val.tzinfo is not None:
            val = val.astimezone(timezone.utc).replace(tzinfo=None)
        return val

    if isinstance(val, time):
        if val.tzinfo is not None:
            val = val.replace(tzinfo=None)
        return val

    if isinstance(val, date):
        return val

    if isinstance(val, Enum):
        return val.value

    if isinstance(val, UUID):
        return str(val)

    if isinstance(val, (bytes, bytearray)):
        return val.decode("utf-8", errors="ignore")

    if isinstance(val, (dict, list, tuple, set)):
        return str(val)

    return str(val)


def _get_material_pairs(g: Goods, max_pairs: int = 10) -> List[Tuple[str, Any]]:
    """
    将 g.materials（material_name/quantity/unit）展开成：
    [(材料1, name1), (材料1用量, qty1), (材料1用量单位, unit1), ...]
    最多展开 max_pairs（默认 10）。
    """
    pairs: List[Tuple[str, Any]] = []
    mats = list(getattr(g, "materials", []) or [])

    for i, m in enumerate(mats[:max_pairs], start=1):
        name = _excel_sanitize_value(getattr(m, "material_name", None))
        qty  = _excel_sanitize_value(getattr(m, "quantity", None))
        unit = _excel_sanitize_value(getattr(m, "unit", None)) or "件"

        pairs.append((f"材料{i}", name))
        pairs.append((f"材料{i}用量", qty))
        pairs.append((f"材料{i}用量单位", unit))

    return pairs


def _goods_to_row(g: Goods) -> Dict[str, Any]:
    """
    将一条 Goods 扁平为按 ORDERED_COLUMNS 对齐的 dict。
    未出现的列统一补 None；“选择框”固定输出 '□'。
    """
    row: Dict[str, Any] = {}

    # 选择框
    row["选择框"] = "□"

    # 销售信息（主表）
    row["大类目"] = _excel_sanitize_value(g.category)
    row["小品类"] = _excel_sanitize_value(g.subcategory)
    row["季节性"] = _excel_sanitize_value(g.season)
    row["产品"] = _excel_sanitize_value(g.product_name)
    row["销售渠道"] = _excel_sanitize_value(g.channel)
    row["责任人"] = _excel_sanitize_value(g.owner)
    row["SKU"] = _excel_sanitize_value(g.sku)
    row["ASIN"] = _excel_sanitize_value(g.asin)
    row["产品条码"] = _excel_sanitize_value(g.barcode)
    row["自定义箱唛"] = _excel_sanitize_value(getattr(g, "carton_mark", None))
    row["货号"] = _excel_sanitize_value(getattr(g, "item_no", None))
    row["颜色"] = _excel_sanitize_value(getattr(g, "color", None))
    row["尺寸"] = _excel_sanitize_value(getattr(g, "size", None))
    row["销售价"] = _excel_sanitize_value(getattr(g, "sale_price", None))

    # 供应信息（一对一）
    s = getattr(g, "supply", None)
    row["供应商"] = _excel_sanitize_value(getattr(s, "vendor", None)) if s else None
    row["采购价"] = _excel_sanitize_value(getattr(s, "purchase_price", None)) if s else None
    row["单品包装尺寸"] = _excel_sanitize_value(getattr(s, "pkg_size", None)) if s else None
    row["单品包装重量"] = _excel_sanitize_value(getattr(s, "pkg_weight", None)) if s else None
    row["装箱系数"] = _excel_sanitize_value(getattr(s, "packing_ratio", None)) if s else None
    row["外箱长"] = _excel_sanitize_value(getattr(s, "carton_l", None)) if s else None
    row["外箱宽"] = _excel_sanitize_value(getattr(s, "carton_w", None)) if s else None
    row["外箱高"] = _excel_sanitize_value(getattr(s, "carton_h", None)) if s else None

    # 报关信息（一对一）
    c = getattr(g, "customs", None)
    row["中文品名"] = _excel_sanitize_value(getattr(c, "name_cn", None)) if c else None
    row["英文品名"] = _excel_sanitize_value(getattr(c, "name_en", None)) if c else None
    row["海关编码"] = _excel_sanitize_value(getattr(c, "hscode", None)) if c else None
    row["申报要素"] = _excel_sanitize_value(getattr(c, "declaration", None)) if c else None
    row["申报价"] = _excel_sanitize_value(getattr(c, "declared_price", None)) if c else None
    row["图片"] = _excel_sanitize_value(getattr(c, "image_note", None)) if c else None

    # 生产配套（最多展开到材料10）
    for k, v in _get_material_pairs(g, max_pairs=10):
        row[k] = v

    # 统一按 ORDERED_COLUMNS 补齐缺列
    for col in ORDERED_COLUMNS:
        row.setdefault(col, None)

    return row


def _rows_to_excel_bytes(rows: List[Dict[str, Any]], sheet_name: str = "Sheet1") -> bytes:
    """
    按 ORDERED_COLUMNS 输出为 Excel（二进制），并统一设置样式：
    - 表头：微软雅黑 10pt 加粗 黑色，居中
    - 数据：微软雅黑 10pt 黑色，居中
    - 所有单元格：四边细边框
    """
    if not rows:
        df = pd.DataFrame(columns=ORDERED_COLUMNS)
    else:
        norm = [{k: r.get(k) for k in ORDERED_COLUMNS} for r in rows]
        df = pd.DataFrame(norm, columns=ORDERED_COLUMNS)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # 取得 worksheet
        ws = writer.sheets[sheet_name]

        # 样式：字体 / 对齐 / 边框
        font_header = Font(name="微软雅黑", size=10, bold=True, color="000000")
        font_body   = Font(name="微软雅黑", size=10, bold=False, color="000000")
        align_center = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="000000")
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

        # 设置表头（第1行）
        for cell in ws[1]:
            cell.font = font_header
            cell.alignment = align_center
            cell.border = border_all

        # 设置数据行（第2行到最后）
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row >= 2:
            for r in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
                for cell in r:
                    cell.font = font_body
                    cell.alignment = align_center
                    cell.border = border_all

    buf.seek(0)
    return buf.getvalue()


def export_from_goods_list(
    goods_list: Iterable[Goods],
    *,
    sheet_name: str = "Sheet1",
) -> bytes:
    """
    已有 Goods 列表直接导出（参考 1111.xlsx）。
    """
    rows = [_goods_to_row(g) for g in goods_list]
    return _rows_to_excel_bytes(rows, sheet_name=sheet_name)
